from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

TOOLS = ["FridayReport.AI", "Oracle Primavera P6", "MS Project Online", "MS Planner", "Smartsheet", "Monday.com", "Asana", "Jira"]

YES = "\u2705"
PARTIAL = "\u26A0\uFE0F"
NO = "\u274C"

GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
CATEGORY_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUMMARY_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FRIDAY_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")

HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
CATEGORY_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E79")
FEATURE_FONT = Font(name="Calibri", size=10)
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="666666")
BOLD_FONT = Font(name="Calibri", size=11, bold=True)
SUMMARY_FONT = Font(name="Calibri", size=11, bold=True, color="1F4E79")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

categories = [
    ("1. Portfolio Management", [
        ("Portfolio creation & grouping", ["Y","Y","Y","N","Y","Y","Y","N"]),
        ("Custom portfolios (cross-portfolio projects)", ["Y","N","N","N","N","N","N","N"]),
        ("Portfolio health scoring (RAG)", ["Y","Y","Y","N","P","Y","Y","N"]),
        ("Portfolio budget tracking", ["Y","Y","Y","N","Y","P","N","N"]),
        ("Strategic alignment / objectives", ["Y","Y","Y","N","N","N","Y","N"]),
        ("Portfolio risk tolerance setting", ["Y","Y","P","N","N","N","N","N"]),
        ("Portfolio-level KPIs / performance metrics", ["Y","Y","Y","N","P","Y","Y","N"]),
        ("AI-powered portfolio risk assessment", ["Y","N","N","N","N","N","N","N"]),
        ("Portfolio manager & business owner roles", ["Y","Y","Y","N","N","N","P","N"]),
        ("Portfolio department assignment", ["Y","Y","Y","N","N","N","N","N"]),
    ]),
    ("2. Project Management", [
        ("Project creation & tracking", ["Y","Y","Y","Y","Y","Y","Y","Y"]),
        ("Project health status (RAG)", ["Y","Y","Y","N","Y","Y","Y","N"]),
        ("Health status history tracking", ["Y","P","N","N","N","N","N","N"]),
        ("Project change logs / audit trail", ["Y","Y","P","N","Y","N","N","Y"]),
        ("Project financials (budget / actual cost)", ["Y","Y","Y","N","Y","P","N","N"]),
        ("Cost items breakdown", ["Y","Y","Y","N","Y","N","N","N"]),
        ("Billable status tracking", ["Y","N","N","N","N","N","N","N"]),
        ("Project scoring / prioritization criteria", ["Y","Y","Y","N","P","P","Y","N"]),
        ("Project benefits tracking", ["Y","P","Y","N","N","N","Y","N"]),
        ("Project decisions log", ["Y","N","N","N","N","N","N","N"]),
        ("Custom fields per project", ["Y","Y","Y","N","Y","Y","Y","Y"]),
        ("Custom project tabs / sections", ["Y","N","N","N","N","N","N","N"]),
        ("Saved project views / filters", ["Y","Y","Y","N","Y","Y","Y","Y"]),
        ("Gantt chart / visual scheduling", ["Y","Y","Y","N","Y","Y","Y","P"]),
        ("Critical path analysis", ["N","Y","Y","N","Y","N","N","N"]),
    ]),
    ("3. Task Management", [
        ("Task creation & assignment", ["Y","Y","Y","Y","Y","Y","Y","Y"]),
        ("Task dependencies (predecessor / successor)", ["Y","Y","Y","N","Y","Y","Y","Y"]),
        ("Hierarchical tasks / WBS roll-up", ["Y","Y","Y","N","Y","Y","Y","Y"]),
        ("Task change logs / audit trail", ["Y","Y","P","N","Y","N","N","Y"]),
        ("Task resource assignments", ["Y","Y","Y","Y","Y","Y","Y","Y"]),
        ("Kanban board view", ["Y","N","N","Y","Y","Y","Y","Y"]),
        ("Recurring tasks", ["N","N","N","Y","Y","Y","Y","N"]),
    ]),
    ("4. Resource Management", [
        ("Resource pool / directory", ["Y","Y","Y","N","Y","Y","Y","N"]),
        ("Resource skills & proficiency levels", ["Y","Y","Y","N","N","N","N","N"]),
        ("Resource availability / time-off calendar", ["Y","Y","Y","N","P","P","N","N"]),
        ("Capacity planning view", ["Y","Y","Y","N","Y","Y","Y","N"]),
        ("Workload dashboard", ["Y","Y","Y","N","Y","Y","Y","N"]),
        ("Demand vs. supply forecast", ["Y","Y","Y","N","N","N","N","N"]),
        ("Resource credit / cost rates", ["Y","Y","Y","N","Y","N","N","N"]),
        ("Resource invitation workflow", ["Y","N","N","N","N","Y","Y","Y"]),
    ]),
    ("5. Time Management", [
        ("Timesheet entries (per task)", ["Y","Y","Y","N","Y","Y","N","P"]),
        ("Timesheet approval periods", ["Y","Y","Y","N","N","N","N","N"]),
        ("Non-project time tracking", ["Y","Y","Y","N","N","N","N","N"]),
        ("Time categories (customizable)", ["Y","Y","P","N","N","N","N","N"]),
        ("Calendar view", ["Y","P","Y","Y","Y","Y","Y","N"]),
    ]),
    ("6. Risk & Issue Management", [
        ("Risk register", ["Y","Y","Y","N","P","N","N","N"]),
        ("AI-powered risk assessment (project level)", ["Y","N","N","N","N","N","N","N"]),
        ("AI-powered risk assessment (portfolio level)", ["Y","N","N","N","N","N","N","N"]),
        ("Risk change logs / history", ["Y","P","N","N","N","N","N","N"]),
        ("Risk resource assignments", ["Y","N","N","N","N","N","N","N"]),
        ("Shareable public risk assessment links", ["Y","N","N","N","N","N","N","N"]),
        ("Issue tracking", ["Y","P","Y","N","Y","Y","Y","Y"]),
        ("Issue resource assignments", ["Y","N","N","N","N","N","Y","Y"]),
        ("Issue change logs / history", ["Y","N","N","N","N","N","N","Y"]),
    ]),
    ("7. Change & Governance", [
        ("Change request management", ["Y","N","N","N","N","N","N","P"]),
        ("Project intake / demand management", ["Y","N","Y","N","P","N","N","N"]),
        ("Intake workflow steps (configurable)", ["Y","N","Y","N","N","N","N","N"]),
        ("Milestone tracking", ["Y","Y","Y","N","Y","Y","Y","Y"]),
        ("Lessons learned module", ["Y","N","N","N","N","N","N","N"]),
        ("Project comments / collaboration", ["Y","P","Y","Y","Y","Y","Y","Y"]),
        ("Project documents management", ["Y","P","Y","N","Y","Y","Y","Y"]),
    ]),
    ("8. Reporting & Analytics", [
        ("Dashboard with charts & visualizations", ["Y","Y","Y","P","Y","Y","Y","Y"]),
        ("Customizable dashboard tab order", ["Y","Y","Y","N","Y","Y","Y","Y"]),
        ("Custom dashboards (user-defined)", ["Y","Y","Y","N","Y","Y","Y","Y"]),
        ("Status report history", ["Y","Y","P","N","Y","N","N","N"]),
        ("Report subscriptions (scheduled email)", ["Y","Y","Y","N","Y","N","Y","N"]),
        ("Analytics API for external tools (Power BI)", ["Y","Y","Y","N","Y","N","N","N"]),
        ("Simulation / what-if analysis", ["Y","Y","Y","N","N","N","N","N"]),
    ]),
    ("9. Financial Management", [
        ("Project invoicing module", ["Y","N","N","N","N","N","N","N"]),
        ("Invoice notes & tracking", ["Y","N","N","N","N","N","N","N"]),
        ("Microsoft Dynamics 365 invoice import", ["Y","N","N","N","N","N","N","N"]),
        ("Project financials module (budget vs. actual)", ["Y","Y","Y","N","P","N","N","N"]),
        ("Cost items breakdown", ["Y","Y","Y","N","Y","N","N","N"]),
        ("Billable status with comments", ["Y","N","N","N","N","N","N","N"]),
    ]),
    ("10. Integrations & Import/Export", [
        ("MS Project file import (.mpp, XML, CSV)", ["Y","Y","Y","N","Y","N","N","N"]),
        ("Microsoft Planner sync", ["Y","N","N","Y","N","N","N","N"]),
        ("Microsoft Dynamics 365 integration", ["Y","N","N","N","N","N","N","N"]),
        ("REST API for external tools", ["Y","Y","Y","P","Y","Y","Y","Y"]),
        ("Swagger / OpenAPI documentation", ["Y","N","P","N","Y","Y","Y","Y"]),
        ("Google Analytics integration", ["Y","N","N","N","N","N","N","N"]),
        ("Zapier / Power Automate", ["N","P","Y","Y","Y","Y","Y","Y"]),
        ("Slack / Teams integration", ["N","N","Y","Y","Y","Y","Y","Y"]),
    ]),
    ("11. Administration & Security", [
        ("Multi-tenant organizations", ["Y","Y","Y","Y","Y","Y","Y","Y"]),
        ("Role-based access control (RBAC)", ["Y","Y","Y","P","Y","Y","Y","Y"]),
        ("Team member scoped access", ["Y","Y","Y","N","Y","P","P","Y"]),
        ("External sharing (read-only viewers)", ["Y","N","P","N","Y","Y","Y","N"]),
        ("Soft delete / data protection", ["Y","P","N","N","N","Y","Y","N"]),
        ("Seat-based billing & plan management", ["Y","Y","Y","Y","Y","Y","Y","Y"]),
        ("Super admin console", ["Y","Y","Y","N","Y","Y","Y","Y"]),
        ("User consent tracking (GDPR compliance)", ["Y","P","Y","Y","N","Y","Y","Y"]),
        ("Email verification", ["Y","Y","Y","Y","N","Y","Y","Y"]),
        ("Bot protection (honeypot + time-based)", ["Y","N","N","N","N","N","N","N"]),
        ("In-app help ticket system", ["Y","N","N","N","N","N","N","N"]),
        ("Notification engine (multi-type, severity)", ["Y","Y","Y","Y","Y","Y","Y","Y"]),
        ("Customizable sidebar navigation", ["Y","N","N","N","N","Y","Y","Y"]),
        ("Organization-scoped integration settings", ["Y","Y","Y","N","Y","Y","Y","Y"]),
    ]),
    ("12. User Experience & Platform", [
        ("Modern web-based UI", ["Y","P","Y","Y","Y","Y","Y","Y"]),
        ("Mobile application", ["N","Y","Y","Y","Y","Y","Y","Y"]),
        ("Dark mode", ["Y","N","N","N","N","Y","Y","Y"]),
        ("User onboarding flow", ["Y","P","P","N","Y","Y","Y","Y"]),
        ("In-app user guide", ["Y","Y","Y","N","Y","Y","Y","Y"]),
        ("Demo data generation", ["Y","N","N","N","Y","Y","N","N"]),
        ("Landing page / public website", ["Y","N","N","N","N","N","N","N"]),
        ("Magic link / passwordless sign-in", ["Y","N","N","N","N","Y","N","N"]),
    ]),
]

def get_display(val):
    if val == "Y": return "Yes"
    if val == "P": return "Partial"
    return "No"

def get_fill(val):
    if val == "Y": return GREEN_FILL
    if val == "P": return YELLOW_FILL
    return RED_FILL

def get_font_color(val):
    if val == "Y": return Font(name="Calibri", size=10, color="006100")
    if val == "P": return Font(name="Calibri", size=10, color="9C6500")
    return Font(name="Calibri", size=10, color="9C0006")

ws = wb.active
ws.title = "Feature Comparison"

ws.merge_cells("A1:I1")
cell = ws["A1"]
cell.value = "FridayReport.AI \u2014 Feature Comparison vs Market Leaders"
cell.font = TITLE_FONT
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 35

ws.merge_cells("A2:I2")
cell = ws["A2"]
cell.value = "Date: February 2026  |  Document Version: 1.0"
cell.font = SUBTITLE_FONT
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 20

ws.merge_cells("A3:I3")
cell = ws["A3"]
cell.value = "Legend:  Yes = Fully supported  |  Partial = Partially supported or requires add-ons  |  No = Not available"
cell.font = Font(name="Calibri", size=9, italic=True, color="444444")
cell.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[3].height = 18

row = 5

for cat_name, features in categories:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    cell = ws.cell(row=row, column=1, value=cat_name)
    cell.font = CATEGORY_FONT
    cell.fill = CATEGORY_FILL
    cell.alignment = LEFT
    cell.border = THIN_BORDER
    for c in range(2, 10):
        ws.cell(row=row, column=c).fill = CATEGORY_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.row_dimensions[row].height = 26
    row += 1

    headers = ["Feature"] + TOOLS
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        if col_idx == 2:
            cell.fill = FRIDAY_FILL
        else:
            cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 30
    row += 1

    for feature_name, vals in features:
        ws.cell(row=row, column=1, value=feature_name).font = FEATURE_FONT
        ws.cell(row=row, column=1).alignment = LEFT
        ws.cell(row=row, column=1).border = THIN_BORDER

        for col_idx, val in enumerate(vals, 2):
            cell = ws.cell(row=row, column=col_idx, value=get_display(val))
            cell.fill = get_fill(val)
            cell.font = get_font_color(val)
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    row += 1

ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
cell = ws.cell(row=row, column=1, value="Feature Count Summary")
cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
cell.fill = CATEGORY_FILL
cell.alignment = LEFT
cell.border = THIN_BORDER
for c in range(2, 10):
    ws.cell(row=row, column=c).fill = CATEGORY_FILL
    ws.cell(row=row, column=c).border = THIN_BORDER
ws.row_dimensions[row].height = 30
row += 1

summary_headers = ["Category"] + TOOLS
for col_idx, h in enumerate(summary_headers, 1):
    cell = ws.cell(row=row, column=col_idx, value=h)
    if col_idx == 2:
        cell.fill = FRIDAY_FILL
    else:
        cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
ws.row_dimensions[row].height = 28
row += 1

for cat_name, features in categories:
    counts = [0] * 8
    total = len(features)
    for _, vals in features:
        for i, v in enumerate(vals):
            if v == "Y":
                counts[i] += 1

    cell = ws.cell(row=row, column=1, value=cat_name)
    cell.font = SUMMARY_FONT
    cell.alignment = LEFT
    cell.fill = SUMMARY_FILL
    cell.border = THIN_BORDER

    for col_idx, count in enumerate(counts, 2):
        cell = ws.cell(row=row, column=col_idx, value=f"{count}/{total}")
        cell.font = BOLD_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        if count == total:
            cell.fill = GREEN_FILL
            cell.font = Font(name="Calibri", size=11, bold=True, color="006100")
        elif count >= total * 0.7:
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        elif count >= total * 0.4:
            cell.fill = YELLOW_FILL
        else:
            cell.fill = RED_FILL
    ws.row_dimensions[row].height = 24
    row += 1

total_counts = [0] * 8
total_features = sum(len(f) for _, f in categories)
for _, features in categories:
    for _, vals in features:
        for i, v in enumerate(vals):
            if v == "Y":
                total_counts[i] += 1

cell = ws.cell(row=row, column=1, value="TOTAL")
cell.font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
cell.alignment = LEFT
cell.fill = CATEGORY_FILL
cell.border = THIN_BORDER

for col_idx, count in enumerate(total_counts, 2):
    cell = ws.cell(row=row, column=col_idx, value=f"{count}/{total_features}")
    cell.font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
    cell.alignment = CENTER
    cell.fill = CATEGORY_FILL
    cell.border = THIN_BORDER
ws.row_dimensions[row].height = 28
row += 2

ws_diff = wb.create_sheet("Key Differentiators")

ws_diff.merge_cells("A1:B1")
cell = ws_diff["A1"]
cell.value = "FridayReport.AI \u2014 Key Differentiators"
cell.font = TITLE_FONT
cell.alignment = Alignment(horizontal="left", vertical="center")
ws_diff.row_dimensions[1].height = 35

r = 3
ws_diff.cell(row=r, column=1, value="Unique Features (not found in any competitor)").font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
ws_diff.merge_cells(f"A{r}:B{r}")
r += 1

for col_idx, h in enumerate(["Feature", "Description"], 1):
    cell = ws_diff.cell(row=r, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
r += 1

unique_features = [
    ("AI-Powered Risk Assessment", "Automated risk analysis at both project and portfolio levels using configurable AI models"),
    ("Shareable Public Risk Assessment Links", "Share risk reports externally without requiring login"),
    ("Project Invoicing Module", "Built-in invoicing with notes, tracking, and Dynamics 365 import"),
    ("Lessons Learned Module", "Structured knowledge capture for organizational learning"),
    ("Change Request Management", "Formal change control process built into the platform"),
    ("Project Decisions Log", "Track and document project decisions for auditability"),
    ("Custom Project Tabs & Sections", "Extend the project detail view with organization-specific content"),
    ("Custom Portfolios", "Create cross-cutting portfolio views that span organizational boundaries"),
    ("Billable Status Tracking", "Track and comment on project billability status"),
    ("In-App Help Ticket System", "Built-in support with screenshot capture"),
]

for feat, desc in unique_features:
    ws_diff.cell(row=r, column=1, value=feat).font = BOLD_FONT
    ws_diff.cell(row=r, column=1).alignment = LEFT
    ws_diff.cell(row=r, column=1).border = THIN_BORDER
    ws_diff.cell(row=r, column=2, value=desc).font = FEATURE_FONT
    ws_diff.cell(row=r, column=2).alignment = LEFT
    ws_diff.cell(row=r, column=2).border = THIN_BORDER
    ws_diff.row_dimensions[r].height = 24
    r += 1

r += 1
ws_diff.cell(row=r, column=1, value="Areas for Future Enhancement").font = Font(name="Calibri", size=13, bold=True, color="1F4E79")
ws_diff.merge_cells(f"A{r}:B{r}")
r += 1

for col_idx, h in enumerate(["Feature Gap", "Available In"], 1):
    cell = ws_diff.cell(row=r, column=col_idx, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER
r += 1

gaps = [
    ("Critical Path Analysis", "Primavera, MS Project, Smartsheet"),
    ("Mobile Application", "All competitors"),
    ("Workflow Automation Engine", "Smartsheet, Monday, Asana"),
    ("Slack / Teams Messaging Integration", "MS Project, Smartsheet, Monday, Asana, Jira"),
    ("Zapier / Power Automate Connector", "MS Project, Smartsheet, Monday, Asana, Jira"),
    ("Recurring Tasks", "Planner, Smartsheet, Monday, Asana"),
]

for feat, avail in gaps:
    ws_diff.cell(row=r, column=1, value=feat).font = BOLD_FONT
    ws_diff.cell(row=r, column=1).alignment = LEFT
    ws_diff.cell(row=r, column=1).border = THIN_BORDER
    ws_diff.cell(row=r, column=1).fill = YELLOW_FILL
    ws_diff.cell(row=r, column=2, value=avail).font = FEATURE_FONT
    ws_diff.cell(row=r, column=2).alignment = LEFT
    ws_diff.cell(row=r, column=2).border = THIN_BORDER
    ws_diff.row_dimensions[r].height = 24
    r += 1

ws.column_dimensions["A"].width = 45
for i in range(2, 10):
    ws.column_dimensions[get_column_letter(i)].width = 18

ws_diff.column_dimensions["A"].width = 40
ws_diff.column_dimensions["B"].width = 65

ws.sheet_view.showGridLines = False
ws_diff.sheet_view.showGridLines = False

wb.save("FridayReport_AI_Feature_Comparison.xlsx")
print("Excel file generated successfully!")
